import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os

# Browse file: 
def browse_file():
    # Open file dialog to select an Excel file
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
    if filename:
        # Clear the entry box and insert the new path
        entry_file.delete(0, tk.END)
        entry_file.insert(0, filename)

# Run hide:
def run_hide_process():
    # 1. Get user inputs
    file_path = entry_file.get()
    sheet_name = entry_sheet.get()
    cells_input = entry_cells.get() 

    # Validation
    if not file_path or not os.path.exists(file_path):
        messagebox.showwarning("Warning", "Please select a valid Excel file.")
        return
    if not sheet_name:
        messagebox.showwarning("Warning", "Please enter a Sheet Name.")
        return

    try:
        # 2. Loading (Keeps formatting)
        wb = openpyxl.load_workbook(file_path)
        
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            messagebox.showerror("Error", f"Sheet '{sheet_name}' not found in the file.")
            return
        
        ws = wb[sheet_name]

        # 3. Process the cells
        # Split into a list ["B4", "B5"]
        cell_list = cells_input.split(',')
        
        for cell_address in cell_list:
            cell_address = cell_address.strip() # Remove spaces
            if cell_address:
                # DIRECTLY set the cell to None (Empty)
                # This solves the logic issue: we target exactly B4, so B6 is never touched.
                ws[cell_address].value = None

        # 4. Save the file
        folder = os.path.dirname(file_path)
        name = os.path.basename(file_path)
        new_name = "For Translation_" + name
        output_path = os.path.join(folder, new_name)

        wb.save(output_path)

        messagebox.showinfo("Success", f"Done!\n\nCells {cells_input} are now empty.\nSaved to:\n{output_path}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")


# UI
# Create the main window
root = tk.Tk()
root.title("Excel Cell Hider")
root.geometry("500x350")

# 1. File Selection UI
tk.Label(root, text="Step 1: Select Excel File", font=("Arial", 10, "bold")).pack(pady=(15, 5))
entry_file = tk.Entry(root, width=50)
entry_file.pack(padx=10)
tk.Button(root, text="Browse", command=browse_file).pack(pady=5)

# 2. Sheet Name UI
tk.Label(root, text="Step 2: Enter Sheet Name (e.g., Sheet1)", font=("Arial", 10, "bold")).pack(pady=(15, 5))
entry_sheet = tk.Entry(root, width=30)
entry_sheet.pack()
# Set a default value example
entry_sheet.insert(0, "Emotional Function (Recall)")

# 3. Cells to Hide UI
tk.Label(root, text="Step 3: Enter Cells to Hide (comma separated)", font=("Arial", 10, "bold")).pack(pady=(15, 5))
tk.Label(root, text="Example: B4, B5, C10", fg="gray").pack()
entry_cells = tk.Entry(root, width=30)
entry_cells.pack()
entry_cells.insert(0, "B4, B5")

# 4. Run Button
tk.Button(root, text="Run & Save", command=run_hide_process, bg="green", fg="white", font=("Arial", 12, "bold")).pack(pady=30, ipadx=20)

# Start the loop
root.mainloop()